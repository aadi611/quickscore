"""
Document processing service for extracting content from various file types.
"""
import io
import re
import logging
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

import pandas as pd
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Service for processing startup documents including pitch decks and financial data."""

    def __init__(self):
        self.supported_formats = {
            'pdf': self._process_pdf,
            'docx': self._process_docx,
            'xlsx': self._process_excel,
            'csv': self._process_csv
        }

    async def process_document(
        self, 
        file_content: bytes, 
        filename: str, 
        document_type: str = "pitch_deck"
    ) -> Dict[str, Any]:
        """
        Process a document and extract relevant content.
        
        Args:
            file_content: Raw file bytes
            filename: Name of the file
            document_type: Type of document (pitch_deck, financial, other)
            
        Returns:
            Dictionary containing extracted and structured content
        """
        try:
            file_ext = Path(filename).suffix.lower().lstrip('.')
            
            if file_ext not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Process based on file type
            processor = self.supported_formats[file_ext]
            raw_content = processor(file_content)
            
            # Structure content based on document type
            if document_type == "pitch_deck":
                structured_content = self._structure_pitch_deck(raw_content, file_ext)
            elif document_type == "financial":
                structured_content = self._structure_financial_data(raw_content, file_ext)
            else:
                structured_content = self._structure_generic_document(raw_content)
            
            return {
                "success": True,
                "document_type": document_type,
                "file_format": file_ext,
                "raw_content": raw_content,
                "structured_content": structured_content,
                "metadata": {
                    "filename": filename,
                    "content_length": len(str(raw_content)),
                    "extraction_method": file_ext
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing document {filename}: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "document_type": document_type,
                "file_format": file_ext if 'file_ext' in locals() else "unknown"
            }

    def _process_pdf(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text content from PDF."""
        try:
            pdf_reader = PdfReader(io.BytesIO(file_content))
            
            pages_content = []
            total_text = ""
            
            for i, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                pages_content.append({
                    "page_number": i + 1,
                    "text": page_text,
                    "word_count": len(page_text.split())
                })
                total_text += page_text + "\n"
            
            return {
                "total_pages": len(pdf_reader.pages),
                "pages": pages_content,
                "full_text": total_text,
                "word_count": len(total_text.split())
            }
            
        except Exception as e:
            raise Exception(f"PDF processing error: {str(e)}")

    def _process_docx(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text content from DOCX."""
        try:
            doc = Document(io.BytesIO(file_content))
            
            paragraphs = []
            full_text = ""
            
            for i, paragraph in enumerate(doc.paragraphs):
                if paragraph.text.strip():
                    paragraphs.append({
                        "paragraph_number": i + 1,
                        "text": paragraph.text,
                        "style": paragraph.style.name if paragraph.style else "Normal"
                    })
                    full_text += paragraph.text + "\n"
            
            return {
                "total_paragraphs": len(paragraphs),
                "paragraphs": paragraphs,
                "full_text": full_text,
                "word_count": len(full_text.split())
            }
            
        except Exception as e:
            raise Exception(f"DOCX processing error: {str(e)}")

    def _process_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Extract data from Excel files."""
        try:
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Convert to list of lists
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        sheet_data.append(list(row))
                
                # Try to convert to DataFrame for better structure
                try:
                    if sheet_data and len(sheet_data) > 1:
                        df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
                        sheets_data[sheet_name] = {
                            "dataframe": df.to_dict('records'),
                            "columns": list(df.columns),
                            "shape": df.shape,
                            "raw_data": sheet_data
                        }
                    else:
                        sheets_data[sheet_name] = {
                            "raw_data": sheet_data,
                            "columns": [],
                            "shape": (len(sheet_data), len(sheet_data[0]) if sheet_data else 0)
                        }
                except Exception:
                    sheets_data[sheet_name] = {"raw_data": sheet_data}
            
            return {
                "total_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "sheets": sheets_data
            }
            
        except Exception as e:
            raise Exception(f"Excel processing error: {str(e)}")

    def _process_csv(self, file_content: bytes) -> Dict[str, Any]:
        """Extract data from CSV files."""
        try:
            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    csv_text = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise Exception("Unable to decode CSV file")
            
            # Try different delimiters
            for delimiter in [',', ';', '\t']:
                try:
                    df = pd.read_csv(io.StringIO(csv_text), delimiter=delimiter)
                    if df.shape[1] > 1:  # Valid separation found
                        break
                except Exception:
                    continue
            else:
                # Fallback to comma delimiter
                df = pd.read_csv(io.StringIO(csv_text))
            
            return {
                "dataframe": df.to_dict('records'),
                "columns": list(df.columns),
                "shape": df.shape,
                "sample_rows": df.head().to_dict('records') if len(df) > 0 else []
            }
            
        except Exception as e:
            raise Exception(f"CSV processing error: {str(e)}")

    def _structure_pitch_deck(self, raw_content: Dict, file_format: str) -> Dict[str, Any]:
        """Structure pitch deck content into meaningful sections."""
        if file_format == 'pdf':
            text = raw_content.get('full_text', '')
        elif file_format == 'docx':
            text = raw_content.get('full_text', '')
        else:
            return {"error": "Pitch deck analysis not supported for this format"}
        
        # Define section patterns (case-insensitive)
        section_patterns = {
            "problem": [
                r"problem", r"pain point", r"challenge", r"issue", r"difficulty"
            ],
            "solution": [
                r"solution", r"approach", r"how we solve", r"our answer"
            ],
            "market": [
                r"market", r"tam", r"total addressable market", r"market size", 
                r"target market", r"addressable market"
            ],
            "business_model": [
                r"business model", r"revenue model", r"monetization", r"pricing", 
                r"how we make money"
            ],
            "team": [
                r"team", r"founders", r"about us", r"who we are", r"leadership"
            ],
            "traction": [
                r"traction", r"metrics", r"growth", r"progress", r"milestones", 
                r"achievements", r"results"
            ],
            "competition": [
                r"competition", r"competitors", r"competitive", r"alternatives"
            ],
            "financials": [
                r"financials", r"projections", r"forecast", r"revenue", r"funding"
            ],
            "ask": [
                r"ask", r"funding", r"investment", r"raising", r"round", r"capital"
            ]
        }
        
        # Extract sections
        sections = {}
        text_lower = text.lower()
        
        for section_name, patterns in section_patterns.items():
            section_content = []
            
            for pattern in patterns:
                # Find sentences containing the pattern
                sentences = re.split(r'[.!?]+', text)
                for sentence in sentences:
                    if re.search(pattern, sentence.lower()) and len(sentence.strip()) > 20:
                        section_content.append(sentence.strip())
            
            # Remove duplicates and get best matches
            section_content = list(set(section_content))[:3]  # Top 3 matches
            
            if section_content:
                sections[section_name] = {
                    "content": section_content,
                    "confidence": len(section_content) / 3.0  # Confidence based on matches found
                }
        
        # Extract key metrics if present
        metrics = self._extract_metrics(text)
        
        return {
            "sections": sections,
            "key_metrics": metrics,
            "slide_count": raw_content.get('total_pages', 0) if file_format == 'pdf' else 1,
            "total_words": raw_content.get('word_count', 0),
            "completeness_score": len(sections) / len(section_patterns)  # How complete is the deck
        }

    def _structure_financial_data(self, raw_content: Dict, file_format: str) -> Dict[str, Any]:
        """Structure financial document content."""
        if file_format in ['xlsx', 'csv']:
            return self._analyze_financial_spreadsheet(raw_content)
        else:
            # For PDF/DOCX financial documents
            text = raw_content.get('full_text', '')
            return self._extract_financial_text(text)

    def _analyze_financial_spreadsheet(self, raw_content: Dict) -> Dict[str, Any]:
        """Analyze financial data from spreadsheets."""
        financial_data = {
            "revenue_projections": [],
            "expenses": [],
            "burn_rate": None,
            "runway": None,
            "key_metrics": {},
            "time_periods": []
        }
        
        try:
            if 'sheets' in raw_content:
                # Excel file
                for sheet_name, sheet_data in raw_content['sheets'].items():
                    if 'dataframe' in sheet_data:
                        df_data = sheet_data['dataframe']
                        columns = sheet_data.get('columns', [])
                        
                        # Look for financial patterns
                        financial_data.update(self._extract_financial_patterns(df_data, columns))
            
            elif 'dataframe' in raw_content:
                # CSV file
                df_data = raw_content['dataframe']
                columns = raw_content.get('columns', [])
                financial_data.update(self._extract_financial_patterns(df_data, columns))
        
        except Exception as e:
            logger.error(f"Error analyzing financial spreadsheet: {e}")
        
        return financial_data

    def _extract_financial_patterns(self, df_data: List[Dict], columns: List[str]) -> Dict[str, Any]:
        """Extract financial patterns from structured data."""
        patterns = {
            "revenue_patterns": ["revenue", "sales", "income", "earnings"],
            "expense_patterns": ["expense", "cost", "spending", "burn"],
            "date_patterns": ["date", "month", "year", "quarter", "period"]
        }
        
        results = {
            "revenue_data": [],
            "expense_data": [],
            "time_periods": []
        }
        
        # Find relevant columns
        revenue_cols = [col for col in columns if any(pattern in col.lower() for pattern in patterns["revenue_patterns"])]
        expense_cols = [col for col in columns if any(pattern in col.lower() for pattern in patterns["expense_patterns"])]
        date_cols = [col for col in columns if any(pattern in col.lower() for pattern in patterns["date_patterns"])]
        
        # Extract data
        for row in df_data:
            if isinstance(row, dict):
                # Extract revenue data
                for col in revenue_cols:
                    if col in row and row[col] is not None:
                        try:
                            value = float(row[col])
                            results["revenue_data"].append({"period": row.get(date_cols[0] if date_cols else "period", "unknown"), "value": value})
                        except (ValueError, TypeError):
                            pass
                
                # Extract expense data
                for col in expense_cols:
                    if col in row and row[col] is not None:
                        try:
                            value = float(row[col])
                            results["expense_data"].append({"period": row.get(date_cols[0] if date_cols else "period", "unknown"), "value": value})
                        except (ValueError, TypeError):
                            pass
        
        return results

    def _extract_financial_text(self, text: str) -> Dict[str, Any]:
        """Extract financial information from text documents."""
        financial_data = {
            "revenue_mentions": [],
            "funding_mentions": [],
            "metrics": {}
        }
        
        # Revenue patterns
        revenue_patterns = [
            r'\$?(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?\s*(?:in\s+)?(?:revenue|sales|income)',
            r'revenue\s+of\s+\$?(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?'
        ]
        
        # Funding patterns
        funding_patterns = [
            r'raising\s+\$?(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?',
            r'seeking\s+\$?(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?',
            r'funding\s+of\s+\$?(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?'
        ]
        
        for pattern in revenue_patterns:
            matches = re.findall(pattern, text.lower())
            financial_data["revenue_mentions"].extend(matches)
        
        for pattern in funding_patterns:
            matches = re.findall(pattern, text.lower())
            financial_data["funding_mentions"].extend(matches)
        
        return financial_data

    def _extract_metrics(self, text: str) -> Dict[str, Any]:
        """Extract key business metrics from text."""
        metrics = {}
        
        # Common metric patterns
        metric_patterns = {
            "users": r'(\d+(?:,\d+)*)\s*(?:users|customers|subscribers)',
            "revenue": r'\$(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?\s*(?:revenue|sales)',
            "growth": r'(\d+(?:\.\d+)?)\s*%\s*(?:growth|increase)',
            "market_size": r'\$(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?\s*(?:market|tam)',
            "funding": r'\$(\d+(?:,\d+)*(?:\.\d+)?)\s*(?:k|m|million|thousand|billion)?\s*(?:funding|investment|round)'
        }
        
        for metric_name, pattern in metric_patterns.items():
            matches = re.findall(pattern, text.lower())
            if matches:
                metrics[metric_name] = matches
        
        return metrics

    def _structure_generic_document(self, raw_content: Dict) -> Dict[str, Any]:
        """Structure generic document content."""
        if 'full_text' in raw_content:
            text = raw_content['full_text']
            return {
                "content_summary": {
                    "word_count": len(text.split()),
                    "character_count": len(text),
                    "paragraph_count": len([p for p in text.split('\n') if p.strip()]),
                },
                "key_phrases": self._extract_key_phrases(text),
                "content_type": self._classify_content_type(text)
            }
        else:
            return {"content_summary": "Unable to extract text content"}

    def _extract_key_phrases(self, text: str, max_phrases: int = 10) -> List[str]:
        """Extract key phrases from text (simplified implementation)."""
        # Remove common words and extract potential key phrases
        words = re.findall(r'\b[A-Za-z]{3,}\b', text.lower())
        word_freq = {}
        
        for word in words:
            if word not in ['the', 'and', 'for', 'are', 'but', 'not', 'you', 'all', 'can', 'had', 'her', 'was', 'one', 'our', 'out', 'day', 'get', 'has', 'him', 'his', 'how', 'its', 'may', 'new', 'now', 'old', 'see', 'two', 'who', 'boy', 'did', 'man', 'run', 'she', 'use', 'way', 'what', 'with']:
                word_freq[word] = word_freq.get(word, 0) + 1
        
        # Return top phrases
        sorted_words = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)
        return [word for word, freq in sorted_words[:max_phrases]]

    def _classify_content_type(self, text: str) -> str:
        """Classify the type of content based on keywords."""
        text_lower = text.lower()
        
        if any(word in text_lower for word in ['pitch', 'investment', 'funding', 'startup']):
            return 'pitch_deck'
        elif any(word in text_lower for word in ['revenue', 'financial', 'projections', 'budget']):
            return 'financial'
        elif any(word in text_lower for word in ['plan', 'strategy', 'business']):
            return 'business_plan'
        else:
            return 'general'